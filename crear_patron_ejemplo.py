"""
crear_patron_ejemplo.py
-----------------------
Script de ayuda: crea un PLANTILLA.xlsx de EJEMPLO dentro de
la carpeta Trabajos_Cotidianos para que el profesor vea cómo debe quedar.

Ejecutar UNA SOLA VEZ para generar la plantilla:
    python crear_patron_ejemplo.py
"""

from pathlib import Path
import sys
import openpyxl  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

# Cargar TEMAS de config.py
sys.path.insert(0, str(Path(__file__).parent))
from calificador.config import TEMAS  # type: ignore

RUTA = Path(__file__).parent / "PLANTILLA.xlsx"

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────
# Hoja _CONFIG
# ─────────────────────────────────────────────
ws_cfg = wb.active
ws_cfg.title = "_CONFIG"

HEADERS = ["hoja_nombre", "hoja_indice", "tema_id", "celdas_evaluar", "peso"]
for col, h in enumerate(HEADERS, 1):
    c = ws_cfg.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(fill_type="solid", fgColor="0056B3")
    c.alignment = Alignment(horizontal="center")

# Ejemplo con 2 hojas
config_data = [
    # hoja_nombre  idx  tema_id               celdas  peso
    ("Datos",       1,  "nombre_hoja",         "-",    1),
    ("Datos",       1,  "color_hoja",          "-",    1),
    ("Datos",       1,  "operaciones_basicas",  "B2:B8", 2),
    ("Datos",       1,  "promedio",             "B10",  1),
    ("Datos",       1,  "max_min",              "B11:B12", 1),
    ("Analisis",    2,  "nombre_hoja",          "-",    1),
    ("Analisis",    2,  "buscarv",              "D2:D20", 3),
    ("Analisis",    2,  "si_simple",            "E2:E20", 2),
    ("Analisis",    2,  "tabla_dinamica",       "-",    3),
    ("Analisis",    2,  "filtros",              "-",    1),
]

for row_idx, fila in enumerate(config_data, 2):
    for col_idx, val in enumerate(fila, 1):
        ws_cfg.cell(row=row_idx, column=col_idx, value=val)

ws_cfg.column_dimensions["A"].width = 16
ws_cfg.column_dimensions["B"].width = 12
ws_cfg.column_dimensions["C"].width = 24
ws_cfg.column_dimensions["D"].width = 16
ws_cfg.column_dimensions["E"].width = 8

# Lista para validación en Columna H
ws_cfg.cell(row=1, column=8, value="Lista_Temas_ID").font = Font(bold=True)
for i, tema in enumerate(TEMAS, start=2):
    ws_cfg.cell(row=i, column=8, value=tema["id"])
ws_cfg.column_dimensions["H"].width = 25

# ─────────────────────────────────────────────
# Hoja "Datos" — ejemplo solucionario
# ─────────────────────────────────────────────
ws_d = wb.create_sheet("Datos")
ws_d.sheet_properties.tabColor = "FF0000"  # Pestaña roja (ejemplo)

# Datos de ejemplo
ws_d["A1"] = "Producto"
ws_d["B1"] = "Precio"
for i in range(2, 9):
    ws_d[f"A{i}"] = f"Producto {i-1}"
    ws_d[f"B{i}"] = (i - 1) * 1500

# Fórmulas esperadas en el patrón (openpyxl requiere los nombres de función en inglés)
ws_d["A10"] = "Promedio"
ws_d["B10"] = "=AVERAGE(B2:B8)"
ws_d["A11"] = "Máximo"
ws_d["B11"] = "=MAX(B2:B8)"
ws_d["A12"] = "Mínimo"
ws_d["B12"] = "=MIN(B2:B8)"

# ─────────────────────────────────────────────
# Hoja "Analisis" — ejemplo solucionario
# ─────────────────────────────────────────────
ws_a = wb.create_sheet("Analisis")

# Tabla de referencia para BUSCARV
ws_a["F1"] = "Código"
ws_a["G1"] = "Nombre"
ws_a["H1"] = "Precio"
ref = [("C01","Camisa",8000),("C02","Pantalón",12000),("C03","Zapatos",25000)]
for r_idx, (cod,nom,pre) in enumerate(ref, 2):
    ws_a.cell(row=r_idx, column=6, value=cod)
    ws_a.cell(row=r_idx, column=7, value=nom)
    ws_a.cell(row=r_idx, column=8, value=pre)

ws_a["A1"] = "Código"
ws_a["B1"] = "Cantidad"
ws_a["C1"] = "Precio Unitario"  # BUSCARV aquí
ws_a["D1"] = "Descuento"        # SI anidado aquí

for r_idx in range(2, 22):
    ws_a.cell(row=r_idx, column=1, value=f"C0{(r_idx % 3) + 1}")
    ws_a.cell(row=r_idx, column=2, value=r_idx * 2)
    ws_a.cell(row=r_idx, column=3, value=f"=VLOOKUP(A{r_idx},$F$2:$H$4,3,0)")
    ws_a.cell(row=r_idx, column=4, value=f'=IF(B{r_idx}>10,IF(B{r_idx}>20,"20%","10%"),"Sin desc.")')

# Filtro automático
ws_a.auto_filter.ref = "A1:D21"

try:
    wb.save(str(RUTA))
    print(f"✅ Patrón de ejemplo creado en:\n   {RUTA}")
    print("\nAhora abre ese archivo y ajusta las fórmulas según tu propio solucionario.")
except PermissionError:
    print(f"\n❌ ERROR: No se pudo guardar '{RUTA.name}'.")
    print("El archivo está abierto actualmente (probablemente en Excel).")
    print("Por favor, ciérralo y vuelve a ejecutar este script.\n")
