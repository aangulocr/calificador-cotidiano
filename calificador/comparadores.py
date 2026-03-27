"""
comparadores.py — Módulos de comparación por tema.

Cada función recibe:
    hoja_est   : Worksheet del estudiante (abierto con data_only=False para leer fórmulas)
    hoja_pat   : Worksheet patrón (abierto con data_only=False)
    meta       : dict con datos extra del tema (celdas_evaluar, etc.)

Devuelve:
    (puntos_obtenidos: int, puntos_posibles: int, observaciones: list[str])
"""

from __future__ import annotations
import re
from openpyxl.worksheet.worksheet import Worksheet


def _celdas_del_rango(hoja: Worksheet, rango: str):
    """Itera celdas de un rango tipo 'A1:B10', 'B2', o varios separados por coma 'A1, C3:D5'."""
    if not rango or rango.strip() == "-":
        return []
    celdas = []
    # Permite separar rangos no adyacentes por comas, ej: "A2:A10, C2:C10"
    partes = [p.strip() for p in rango.split(',')]
    for parte in partes:
        if not parte:
            continue
        try:
            rango_obj = hoja[parte]
            if isinstance(rango_obj, tuple):
                for fila in rango_obj:
                    if hasattr(fila, '__iter__'):
                        celdas.extend(fila)
                    else:
                        celdas.append(fila)
            else:
                # Es una sola celda
                celdas.append(rango_obj)
        except Exception:
            pass  # Ignorar partes mal formadas
    return celdas


def _formula_en_celda(celda) -> str:
    """Devuelve la fórmula (str) o cadena vacía si no hay."""
    val = celda.value
    if isinstance(val, str) and val.startswith("="):
        return val.upper()
    return ""


def _tiene_funcion(formula: str, *nombres: str) -> bool:
    """Comprueba si la fórmula contiene alguno de los nombres de función dados."""
    for nombre in nombres:
        patron = rf"\b{re.escape(nombre.upper())}\s*\("
        if re.search(patron, formula):
            return True
    return False


def _get_color(color_obj) -> str:
    """
    Normaliza un objeto Color de openpyxl a string hex AARRGGBB.
    Devuelve cadena vacía si no se puede determinar.
    """
    if color_obj is None:
        return ""
    try:
        if hasattr(color_obj, 'rgb') and color_obj.rgb:
            return str(color_obj.rgb).upper()
        if hasattr(color_obj, 'value') and color_obj.value:
            return str(color_obj.value).upper()
    except Exception:
        pass
    return ""


_NEGROS_DEFAULT = {"FF000000", "00000000", "FFFFFFFF", ""}


def _get_border_sig(border) -> str:
    """
    Genera una ‘firma’ del borde de una celda: 'estiloL:estiloR:estiloT:estiloB'.
    Devuelve 'none:none:none:none' si no hay borde.
    """
    if border is None:
        return "none:none:none:none"
    def s(lado):
        return getattr(lado, 'border_style', None) or "none"
    return f"{s(border.left)}:{s(border.right)}:{s(border.top)}:{s(border.bottom)}"


_SIN_BORDE = "none:none:none:none"


# ---------------------------------------------------------------------------
# 1. Nombre de hoja
# ---------------------------------------------------------------------------

def eval_nombre_hoja(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    nombre_esp = hoja_pat.title.strip().lower()
    nombre_est = hoja_est.title.strip().lower()
    obs = []
    if nombre_est == nombre_esp:
        return 1, 1, obs
    obs.append(f"Nombre de hoja incorrecto (esperado: '{hoja_pat.title}', encontrado: '{hoja_est.title}')")
    return 0, 1, obs


# ---------------------------------------------------------------------------
# 2. Color de pestaña
# ---------------------------------------------------------------------------

def eval_color_hoja(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    obs = []
    color_pat = None
    color_est = None
    try:
        color_pat = hoja_pat.sheet_properties.tabColor
        color_est = hoja_est.sheet_properties.tabColor
    except Exception:
        pass

    if color_pat is None:
        # El patrón no tiene color → se acepta cualquier cosa
        return 1, 1, obs

    if color_est is None:
        obs.append("Faltó asignar color a la pestaña")
        return 0, 1, obs

    # Comparar rgb (ignorando tinte/tema)
    rgb_pat = getattr(color_pat, 'rgb', None) or getattr(color_pat, 'value', None)
    rgb_est = getattr(color_est, 'rgb', None) or getattr(color_est, 'value', None)

    if str(rgb_pat).upper() == str(rgb_est).upper():
        return 1, 1, obs

    obs.append(f"Color de pestaña incorrecto (esperado: {rgb_pat}, encontrado: {rgb_est})")
    return 0, 1, obs


# ---------------------------------------------------------------------------
# 3. Operaciones básicas  (+, -, *, / o Funciones Equivalentes)
# ---------------------------------------------------------------------------

def _es_operacion_basica(formula: str) -> bool:
    if not formula or not formula.startswith("="):
        return False
    # Tiene un operador matemático básico:
    if bool(re.search(r"[\+\-\*/]", formula)):
        return True
    # O usa una función matemática equivalente:
    if _tiene_funcion(formula, "SUMA", "SUM", "PRODUCTO", "PRODUCT", "COCIENTE", "QUOTIENT"):
        return True
    return False

def eval_operaciones_basicas(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    rango = meta.get("celdas_evaluar", "")
    obs = []
    if not rango or rango.strip() == "-":
        return 1, 1, obs  # no aplica rango → se acepta

    celdas_pat = _celdas_del_rango(hoja_pat, rango)
    total = len(celdas_pat)
    if total == 0:
        return 1, 1, obs

    correctas = 0
    celdas_faltantes = []

    for cp in celdas_pat:
        formula_pat = _formula_en_celda(cp)
        if not formula_pat:
            total -= 1
            continue
        coord = cp.coordinate
        ce = hoja_est[coord]
        formula_est = _formula_en_celda(ce)
        if _es_operacion_basica(formula_est):
            correctas += 1
        else:
            celdas_faltantes.append(coord)

    if celdas_faltantes:
        obs.append(f"Operaciones básicas ausentes en: {', '.join(celdas_faltantes)}")

    if total == 0:
        return 1, 1, obs
    return correctas, total, obs


# ---------------------------------------------------------------------------
# 4-12. Funciones Excel (una función genérica parametrizable)
# ---------------------------------------------------------------------------

def _eval_funcion_generica(
    hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict, *nombres_funcion: str
):
    rango = meta.get("celdas_evaluar", "")
    tema_nombre = meta.get("tema_nombre", ", ".join(nombres_funcion))
    obs = []

    celdas_pat = _celdas_del_rango(hoja_pat, rango) if rango and rango.strip() != "-" else []

    if not celdas_pat:
        # Sin rango: buscar en toda la hoja si el patrón usa la función
        pat_usa_fn = any(
            _tiene_funcion(_formula_en_celda(c), *nombres_funcion)
            for row in hoja_pat.iter_rows()
            for c in row
        )
        if not pat_usa_fn:
            return 1, 1, obs  # patrón no lo requiere
        # Buscar en hoja del estudiante
        est_usa_fn = any(
            _tiene_funcion(_formula_en_celda(c), *nombres_funcion)
            for row in hoja_est.iter_rows()
            for c in row
        )
        if est_usa_fn:
            return 1, 1, obs
        obs.append(f"No se encontró función {tema_nombre} en la hoja")
        return 0, 1, obs

    # Con rango específico
    total = 0
    correctas = 0
    faltantes = []

    for cp in celdas_pat:
        formula_pat = _formula_en_celda(cp)
        if not _tiene_funcion(formula_pat, *nombres_funcion):
            continue
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        formula_est = _formula_en_celda(ce)
        if _tiene_funcion(formula_est, *nombres_funcion):
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        obs.append(f"Función {tema_nombre} ausente en: {', '.join(faltantes)}")
    return correctas, total, obs


def eval_concatenar(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "CONCATENAR", "CONCAT", "CONCATENATE")

def eval_contar(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "CONTAR", "COUNT")

def eval_contara(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "CONTARA", "COUNTA")

def eval_contar_si(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "CONTAR.SI", "COUNTIF")

def eval_contar_blanco(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "CONTAR.BLANCO", "COUNTBLANK")

def eval_promedio(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "PROMEDIO", "AVERAGE")

def eval_mediana(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "MEDIANA", "MEDIAN")

def eval_moda_uno(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "MODA.UNO", "MODE.SNGL", "MODA")

def eval_max_min(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "MAX", "MIN")

def eval_si_simple(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "SI", "IF")

def eval_si_anidado(hoja_est, hoja_pat, meta):
    """SI anidado → se busca la función SI/IF con otra SI anidada dentro."""
    rango = meta.get("celdas_evaluar", "")
    obs = []
    PATRON_ANIDADO = re.compile(r"\bSI\s*\(.*\bSI\s*\(|\bIF\s*\(.*\bIF\s*\(", re.DOTALL)

    def tiene_si_anidado(formula: str) -> bool:
        return bool(PATRON_ANIDADO.search(formula))

    celdas_pat = _celdas_del_rango(hoja_pat, rango) if rango and rango.strip() != "-" else []

    if not celdas_pat:
        pat_usa = any(
            tiene_si_anidado(_formula_en_celda(c))
            for row in hoja_pat.iter_rows() for c in row
        )
        if not pat_usa:
            return 1, 1, obs
        est_usa = any(
            tiene_si_anidado(_formula_en_celda(c))
            for row in hoja_est.iter_rows() for c in row
        )
        if est_usa:
            return 1, 1, obs
        obs.append("No se encontró SI anidado en la hoja")
        return 0, 1, obs

    total = 0
    correctas = 0
    faltantes = []
    for cp in celdas_pat:
        if not tiene_si_anidado(_formula_en_celda(cp)):
            continue
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        if tiene_si_anidado(_formula_en_celda(ce)):
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        obs.append(f"SI anidado ausente en: {', '.join(faltantes)}")
    return correctas, total, obs


def eval_buscarv(hoja_est, hoja_pat, meta):
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "BUSCARV", "VLOOKUP")


# ---------------------------------------------------------------------------
# 14. Tabla dinámica
# ---------------------------------------------------------------------------

def eval_tabla_dinamica(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    obs = []
    pivots_pat = getattr(hoja_pat, '_pivots', [])
    if not pivots_pat:
        return 1, 1, obs  # patrón no requiere tabla dinámica

    pivots_est = getattr(hoja_est, '_pivots', [])
    if pivots_est:
        return 1, 1, obs

    obs.append("Faltó crear tabla dinámica en la hoja")
    return 0, 1, obs


# ---------------------------------------------------------------------------
# 15. Gráfico dinámico
# ---------------------------------------------------------------------------

def eval_grafico_dinamico(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    obs = []
    charts_pat = getattr(hoja_pat, '_charts', [])
    if not charts_pat:
        return 1, 1, obs

    charts_est = getattr(hoja_est, '_charts', [])
    if charts_est:
        return 1, 1, obs

    obs.append("Faltó crear gráfico dinámico en la hoja")
    return 0, 1, obs


# ---------------------------------------------------------------------------
# 16. Gráfico normal
# ---------------------------------------------------------------------------

def eval_grafico_normal(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    obs = []
    charts_pat = getattr(hoja_pat, '_charts', [])
    if not charts_pat:
        return 1, 1, obs

    charts_est = getattr(hoja_est, '_charts', [])
    if charts_est:
        return 1, 1, obs

    obs.append("Faltó insertar gráfico en la hoja")
    return 0, 1, obs


# ---------------------------------------------------------------------------
# 17. Filtros automáticos
# ---------------------------------------------------------------------------

def eval_filtros(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    obs = []
    tiene_filtro_pat = bool(hoja_pat.auto_filter and hoja_pat.auto_filter.ref)
    if not tiene_filtro_pat:
        return 1, 1, obs

    tiene_filtro_est = bool(hoja_est.auto_filter and hoja_est.auto_filter.ref)
    if tiene_filtro_est:
        return 1, 1, obs

    obs.append("Faltó activar filtros automáticos en la hoja")
    return 0, 1, obs


# ---------------------------------------------------------------------------
# 18. Formato condicional
# ---------------------------------------------------------------------------

def eval_formato_condicional(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    obs = []
    reglas_pat = list(hoja_pat.conditional_formatting)
    if not reglas_pat:
        return 1, 1, obs

    reglas_est = list(hoja_est.conditional_formatting)
    if reglas_est:
        return 1, 1, obs

    obs.append("Faltó aplicar formato condicional en la hoja")
    return 0, 1, obs


# ---------------------------------------------------------------------------
# 19. Validación de datos
# ---------------------------------------------------------------------------

def eval_validacion_datos(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    obs = []
    dv_pat = hoja_pat.data_validations.dataValidation if hoja_pat.data_validations else []
    if not list(dv_pat):
        return 1, 1, obs

    dv_est = hoja_est.data_validations.dataValidation if hoja_est.data_validations else []
    if list(dv_est):
        return 1, 1, obs

    obs.append("Faltó aplicar validación de datos en la hoja")
    return 0, 1, obs


# ---------------------------------------------------------------------------
# 20. Bordes de celdas
# ---------------------------------------------------------------------------

def eval_bordes(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """
    Compara los estilos de borde (left, right, top, bottom) de cada celda
    del rango indicado en el patrón.
    Solo evalúa celdas donde el patrón tiene borde definido (no 'none').
    """
    rango = meta.get("celdas_evaluar", "")
    obs = []
    if not rango or rango.strip() == "-":
        return 1, 1, obs

    celdas_pat = _celdas_del_rango(hoja_pat, rango)
    total = 0
    correctas = 0
    faltantes = []

    for cp in celdas_pat:
        sig_pat = _get_border_sig(cp.border)
        if sig_pat == _SIN_BORDE:
            continue          # patrón no tiene borde aquí → no se evalúa
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        if _get_border_sig(ce.border) == sig_pat:
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        muestra = faltantes[:5]
        extra   = f" (+{len(faltantes)-5} más)" if len(faltantes) > 5 else ""
        obs.append(f"Bordes incorrectos en: {', '.join(muestra)}{extra}")
    return correctas, total, obs


# ---------------------------------------------------------------------------
# 21. Relleno / Color de fondo
# ---------------------------------------------------------------------------

def eval_relleno(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """
    Compara el color de relleno (fgColor) de cada celda del rango.
    Omite celdas sin relleno explícito en el patrón y las que tienen
    colores estándar (blanco/sin color).
    """
    rango = meta.get("celdas_evaluar", "")
    obs = []
    if not rango or rango.strip() == "-":
        return 1, 1, obs

    celdas_pat = _celdas_del_rango(hoja_pat, rango)
    total = 0
    correctas = 0
    faltantes = []

    for cp in celdas_pat:
        fill_pat = cp.fill
        if fill_pat is None or fill_pat.patternType in (None, "none"):
            continue
        color_pat = _get_color(getattr(fill_pat, 'fgColor', None))
        if not color_pat or color_pat in _NEGROS_DEFAULT:
            continue          # color por defecto → no se evalúa
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        color_est = _get_color(getattr(ce.fill, 'fgColor', None)) if ce.fill else ""
        if color_est == color_pat:
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        muestra = faltantes[:5]
        extra   = f" (+{len(faltantes)-5} más)" if len(faltantes) > 5 else ""
        obs.append(f"Relleno incorrecto en: {', '.join(muestra)}{extra}")
    return correctas, total, obs


# ---------------------------------------------------------------------------
# 22. Color de fuente
# ---------------------------------------------------------------------------

def eval_color_fuente(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """
    Compara el color de la fuente de cada celda del rango.
    Omite negro por defecto (FF000000) para no penalizar texto sin formato.
    """
    rango = meta.get("celdas_evaluar", "")
    obs = []
    if not rango or rango.strip() == "-":
        return 1, 1, obs

    celdas_pat = _celdas_del_rango(hoja_pat, rango)
    total = 0
    correctas = 0
    faltantes = []

    for cp in celdas_pat:
        font_pat = cp.font
        if font_pat is None:
            continue
        color_pat = _get_color(getattr(font_pat, 'color', None))
        if not color_pat or color_pat in _NEGROS_DEFAULT:
            continue          # negro (por defecto) → no se evalúa
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        color_est = _get_color(getattr(ce.font, 'color', None)) if ce.font else ""
        if color_est == color_pat:
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        muestra = faltantes[:5]
        extra   = f" (+{len(faltantes)-5} más)" if len(faltantes) > 5 else ""
        obs.append(f"Color de fuente incorrecto en: {', '.join(muestra)}{extra}")
    return correctas, total, obs


# ---------------------------------------------------------------------------
# Helpers para clasificación de formato numérico
# ---------------------------------------------------------------------------

import re as _re

# Tokens que indican formato de FECHA (patrones de fecha/hora de Excel)
# Tokens que indican formato de FECHA (patrones de fecha/hora de Excel)
_FECHA_TOKENS = _re.compile(
    r'(?<!\\")'          # no escapado
    r"(?:d{1,4}|m{1,5}|y{2,4}|h{1,2}|s{1,2})"  # d, dd, ddd, dddd, m,..., yy, yyyy, h, hh, ss
    r'(?!\\")' ,
    _re.IGNORECASE,
)

# Patrones que indican formato de MONEDA
_MONEDA_TOKENS = _re.compile(
    r'[\$€£₡¥]'          # símbolos de moneda Unicode
    r'|"\$"'                  # $  entre comillas en el código de formato
    r'|#,##0'                 # separador de miles típico de moneda
    r'|_\('                   # formato contable con sangría
    r'|\[\$',                 # código de moneda local [\$-nnn]
    _re.IGNORECASE,
)


def _clasificar_formato(fmt: str) -> str:
    """
    Clasifica el código de formato de Excel en una categoría simplificada.
    Devuelve: 'fecha', 'moneda', 'otro', o 'general'.
    """
    if not fmt or fmt.upper() in ("GENERAL", "@", ""):
        return "general"
    if _FECHA_TOKENS.search(fmt):
        return "fecha"
    if _MONEDA_TOKENS.search(fmt):
        return "moneda"
    return "otro"


def _evaluar_formato_categorico(
    hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict, categoria: str, etiqueta: str
):
    """
    Función genérica: verifica que las celdas del rango tengan
    la misma categoría de formato que el patrón (fecha o moneda).
    """
    rango = meta.get("celdas_evaluar", "")
    obs = []

    celdas_pat = _celdas_del_rango(hoja_pat, rango) if rango and rango.strip() != "-" else []

    if not celdas_pat:
        # Sin rango: buscar en toda la hoja del patrón
        tiene_pat = any(
            _clasificar_formato(c.number_format) == categoria
            for row in hoja_pat.iter_rows() for c in row
        )
        if not tiene_pat:
            return 1, 1, obs   # el patrón no lo requiere
        tiene_est = any(
            _clasificar_formato(c.number_format) == categoria
            for row in hoja_est.iter_rows() for c in row
        )
        if tiene_est:
            return 1, 1, obs
        obs.append(f"No se encontró formato de {etiqueta} en la hoja")
        return 0, 1, obs

    # Con rango específico: evaluar celda por celda
    total = 0
    correctas = 0
    faltantes = []

    for cp in celdas_pat:
        if _clasificar_formato(cp.number_format) != categoria:
            continue           # el patrón no exige ese formato aquí
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        if _clasificar_formato(ce.number_format) == categoria:
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        muestra = faltantes[:5]
        extra   = f" (+{len(faltantes)-5} más)" if len(faltantes) > 5 else ""
        obs.append(f"Formato de {etiqueta} ausente en: {', '.join(muestra)}{extra}")
    return correctas, total, obs


# ---------------------------------------------------------------------------
# 23. Formato de moneda
# ---------------------------------------------------------------------------

def eval_formato_moneda(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """
    Verifica que las celdas del rango tengan el mismo formato de moneda.
    Se asegura que el símbolo explícito usado por el patrón ($ o ₡) coincida
    estrictamente en el archivo del estudiante.
    """
    rango = meta.get("celdas_evaluar", "")
    obs = []
    
    celdas_pat = _celdas_del_rango(hoja_pat, rango) if rango and rango.strip() != "-" else []
    
    def extraer_moneda(fmt: str) -> str:
        if not fmt: return ""
        match = _re.search(r'[\$€£₡¥]', fmt)
        if match: return match.group(0)
        # Si tiene la palabra contable, o guiones, pero no símbolo, 
        # devolvemos "moneda_generica"
        if _MONEDA_TOKENS.search(fmt): return "moneda_generica"
        return ""

    if not celdas_pat:
        return 1, 1, obs # si no hay rango, se da por bueno

    total = 0
    correctas = 0
    faltantes = []

    for cp in celdas_pat:
        simbolo_pat = extraer_moneda(cp.number_format)
        if not simbolo_pat:
            continue
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        
        simbolo_est = extraer_moneda(ce.number_format)
        
        if simbolo_est == simbolo_pat:
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        muestra = faltantes[:5]
        extra   = f" (+{len(faltantes)-5} más)" if len(faltantes) > 5 else ""
        obs.append(f"Moneda incorrecta (simbolo distinto) en: {', '.join(muestra)}{extra}")
    return correctas, total, obs


# ---------------------------------------------------------------------------
# 24. Formato de fecha
# ---------------------------------------------------------------------------

def eval_formato_fecha(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """
    Verifica que las celdas del rango tengan formato de fecha aplicado.
    Detecta tokens: d, dd, ddd, dddd, m, mm, yy, yyyy, h, hh, ss, etc.
    Acepta variantes (dd/mm/yyyy, d-m-yy, yyyy-mm-dd, etc.).
    """
    return _evaluar_formato_categorico(hoja_est, hoja_pat, meta, "fecha", "fecha")



# ---------------------------------------------------------------------------
# 25. SUMAR.SI
# ---------------------------------------------------------------------------

def eval_sumar_si(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """Detecta el uso de SUMAR.SI / SUMIF."""
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "SUMAR.SI", "SUMIF")


# ---------------------------------------------------------------------------
# 26. SI.CONJUNTO / IFS
# ---------------------------------------------------------------------------

def eval_si_conjunto(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """Detecta el uso de SI.CONJUNTO / IFS."""
    return _eval_funcion_generica(hoja_est, hoja_pat, meta, "SI.CONJUNTO", "IFS")


# ---------------------------------------------------------------------------
# 27. Operaciones combinadas (regla de tres, combinaciones aritméticas)
# ---------------------------------------------------------------------------

# Patrón: la fórmula tiene al menos 2 referencias de celda y al menos un operador
# aritmético entre ellas → cubre regla de tres (=A2*B2/C2) y otras combinaciones.
_RE_CELDA_REF   = re.compile(r"\$?[A-Z]{1,3}\$?\d+")        # A1, $B$2, AB12…
_RE_OP_BASICO   = re.compile(r"[\+\-\*/]")                   # cualquier operador


def _es_operacion_combinada(formula: str) -> bool:
    """True si la fórmula tiene ≥2 referencias de celda + ≥1 operador aritmético."""
    if not formula or not formula.startswith("="):
        return False
    refs = _RE_CELDA_REF.findall(formula)
    return len(refs) >= 2 and bool(_RE_OP_BASICO.search(formula[1:]))  # quitar '='


def eval_operaciones_combinadas(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """
    Detecta fórmulas con operadores aritméticos combinando múltiples referencias
    de celda (regla de tres, cadenas de cálculo, etc.).
    No se activa cuando la única "operación" es dentro de una función estándar
    del patrón que ya tiene su propio evaluador.
    """
    rango = meta.get("celdas_evaluar", "")
    obs = []

    celdas_pat = _celdas_del_rango(hoja_pat, rango) if rango and rango.strip() != "-" else []

    if not celdas_pat:
        pat_usa = any(
            _es_operacion_combinada(_formula_en_celda(c))
            for row in hoja_pat.iter_rows() for c in row
        )
        if not pat_usa:
            return 1, 1, obs
        est_usa = any(
            _es_operacion_combinada(_formula_en_celda(c))
            for row in hoja_est.iter_rows() for c in row
        )
        if est_usa:
            return 1, 1, obs
        obs.append("No se encontró operación combinada (regla de tres u otro) en la hoja")
        return 0, 1, obs

    total = 0
    correctas = 0
    faltantes = []
    for cp in celdas_pat:
        if not _es_operacion_combinada(_formula_en_celda(cp)):
            continue
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        if _es_operacion_combinada(_formula_en_celda(ce)):
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        obs.append(f"Operación combinada ausente en: {', '.join(faltantes)}")
    return correctas, total, obs


# ---------------------------------------------------------------------------
# 28. Cálculo de porcentaje
#     Formas aceptadas:
#       a) Multiplicación por decimal  : =A2*0.13  o  =A2*0.03
#       b) Multiplicación por %        : =A2*13%   o  =A2*3%
#       c) Regla de tres para porcentaje: patrón con *100 o /100 + referencias
# ---------------------------------------------------------------------------

# a) decimal: *0.xx o /0.xx o *.xx (multiplicación o división por decimales)
_RE_POR_DECIMAL  = re.compile(r"[\*/]\s*\d*\.\d+")

# b) porcentaje directo: *xx% o /xx%
_RE_POR_PORCIENTO = re.compile(r"[\*/]\s*\d+(?:\.\d+)?%")

# c) regla de tres: incluye dividir o multiplicar por 100 en cualquier parte de la fórmula
_RE_POR_100      = re.compile(r"[\*/]\s*100\b")


def _es_calculo_porcentaje(formula: str) -> bool:
    """True si la fórmula usa cualquiera de los tres patrones de porcentaje."""
    if not formula or not formula.startswith("="):
        return False
    f = formula[1:]   # quitar '='
    return (
        bool(_RE_POR_DECIMAL.search(f))
        or bool(_RE_POR_PORCIENTO.search(f))
        or bool(_RE_POR_100.search(f))
    )


def eval_calculo_porcentaje(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """
    Detecta cálculos de porcentaje / impuesto / descuento en cualquiera
    de las tres variantes: decimal, % directo o regla de tres.
    """
    rango = meta.get("celdas_evaluar", "")
    obs = []

    celdas_pat = _celdas_del_rango(hoja_pat, rango) if rango and rango.strip() != "-" else []

    if not celdas_pat:
        pat_usa = any(
            _es_calculo_porcentaje(_formula_en_celda(c))
            for row in hoja_pat.iter_rows() for c in row
        )
        if not pat_usa:
            return 1, 1, obs
        est_usa = any(
            _es_calculo_porcentaje(_formula_en_celda(c))
            for row in hoja_est.iter_rows() for c in row
        )
        if est_usa:
            return 1, 1, obs
        obs.append("No se encontró cálculo de porcentaje/impuesto/descuento en la hoja")
        return 0, 1, obs

    total = 0
    correctas = 0
    faltantes = []
    for cp in celdas_pat:
        if not _es_calculo_porcentaje(_formula_en_celda(cp)):
            continue
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        if _es_calculo_porcentaje(_formula_en_celda(ce)):
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        obs.append(f"Cálculo de porcentaje ausente en: {', '.join(faltantes)}")
    return correctas, total, obs


# ---------------------------------------------------------------------------
# 29. SI / SI.CONJUNTO con cálculo matemático en sus argumentos
# ---------------------------------------------------------------------------

# Patrón genérico: función SI o SI.CONJUNTO con operador aritmético o porcentaje
# dentro de la llamada. Ej: =SI(A1*B1>100,"Sí","No")  o  =SI.CONJUNTO(A2*0.13>50,"Alto",...)
_RE_SI_INICIO = re.compile(
    r"\b(SI\.CONJUNTO|IFS|SI|IF)\s*\(",
    re.IGNORECASE,
)


def _es_si_con_calculo(formula: str) -> bool:
    """
    Detecta SI / SI.CONJUNTO / IFS cuyo interior contenga
    una operación aritmética o un patrón de porcentaje.
    """
    if not formula or not formula.startswith("="):
        return False
    if not _RE_SI_INICIO.search(formula):
        return False
    # Tomar el contenido después del primer paréntesis de apertura
    m = _RE_SI_INICIO.search(formula)
    interior = formula[m.end():]  # todo lo que va dentro del SI(
    return (
        _es_calculo_porcentaje("=" + interior)
        or _es_operacion_combinada("=" + interior)
    )


def eval_si_con_calculo(hoja_est: Worksheet, hoja_pat: Worksheet, meta: dict):
    """
    Detecta el uso de SI o SI.CONJUNTO / IFS donde alguno de sus
    argumentos contiene una operación combinada o un cálculo de porcentaje.
    Ej: =SI(A1*B1>500,"Aprobado","Reprobado")
        =SI.CONJUNTO(A2*0.13>=100,"Alto",A2*0.13>=50,"Medio",VERDADERO,"Bajo")
    """
    rango = meta.get("celdas_evaluar", "")
    obs = []

    celdas_pat = _celdas_del_rango(hoja_pat, rango) if rango and rango.strip() != "-" else []

    if not celdas_pat:
        pat_usa = any(
            _es_si_con_calculo(_formula_en_celda(c))
            for row in hoja_pat.iter_rows() for c in row
        )
        if not pat_usa:
            return 1, 1, obs
        est_usa = any(
            _es_si_con_calculo(_formula_en_celda(c))
            for row in hoja_est.iter_rows() for c in row
        )
        if est_usa:
            return 1, 1, obs
        obs.append("No se encontró SI/SI.CONJUNTO con cálculo matemático en la hoja")
        return 0, 1, obs

    total = 0
    correctas = 0
    faltantes = []
    for cp in celdas_pat:
        if not _es_si_con_calculo(_formula_en_celda(cp)):
            continue
        total += 1
        coord = cp.coordinate
        ce = hoja_est[coord]
        if _es_si_con_calculo(_formula_en_celda(ce)):
            correctas += 1
        else:
            faltantes.append(coord)

    if total == 0:
        return 1, 1, obs
    if faltantes:
        obs.append(f"SI/SI.CONJUNTO con cálculo ausente en: {', '.join(faltantes)}")
    return correctas, total, obs


EVALUADORES = {
    "nombre_hoja":         eval_nombre_hoja,
    "color_hoja":          eval_color_hoja,
    "operaciones_basicas": eval_operaciones_basicas,
    "concatenar":          eval_concatenar,
    "contar":              eval_contar,
    "contara":             eval_contara,
    "contar_si":              eval_contar_si,
    "contar_blanco":          eval_contar_blanco,
    "promedio":               eval_promedio,
    "mediana":                eval_mediana,
    "moda_uno":               eval_moda_uno,
    "max_min":                eval_max_min,
    "si_simple":              eval_si_simple,
    "si_anidado":             eval_si_anidado,
    "buscarv":                eval_buscarv,
    # --- Nuevas funciones ---
    "sumar_si":               eval_sumar_si,
    "si_conjunto":            eval_si_conjunto,
    "operaciones_combinadas": eval_operaciones_combinadas,
    "calculo_porcentaje":     eval_calculo_porcentaje,
    "si_con_calculo":         eval_si_con_calculo,
    # --- Análisis y visualización ---
    "tabla_dinamica":         eval_tabla_dinamica,
    "grafico_dinamico":       eval_grafico_dinamico,
    "grafico_normal":         eval_grafico_normal,
    "filtros":                eval_filtros,
    "formato_condicional":    eval_formato_condicional,
    "validacion_datos":       eval_validacion_datos,
    # --- Formato ---
    "bordes":                 eval_bordes,
    "relleno":                eval_relleno,
    "color_fuente":           eval_color_fuente,
    "formato_moneda":         eval_formato_moneda,
    "formato_fecha":          eval_formato_fecha,
}
