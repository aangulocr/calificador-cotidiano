"""
exportador.py — Genera Resultados_Evaluacion.xlsx con formato profesional.
"""

from __future__ import annotations
from pathlib import Path
from datetime import date

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

from .config import (
    COLOR_PRIMARIO, ARCHIVO_RESULTADOS, RUTA_RESULTADOS,
    COLOR_LOGRADO, COLOR_PROCESO, COLOR_NO_LOGO
)


# ---------------------------------------------------------------------------
# Paleta interna (hex sin #)
# ---------------------------------------------------------------------------
_AZUL    = "0056B3"
_BLANCO  = "FFFFFF"
_GRIS    = "F5F5F5"
_VERDE   = "198754"
_AMARIL  = "FFC107"
_ROJO    = "DC3545"
_GRIS_B  = "DEE2E6"


def _lado(color="AAAAAA", estilo="thin"):
    return Side(border_style=estilo, color=color)


BORDE_NORMAL   = Border(left=_lado(), right=_lado(), top=_lado(), bottom=_lado())
BORDE_CABECERA = Border(
    left=_lado(_AZUL, "medium"), right=_lado(_AZUL, "medium"),
    top=_lado(_AZUL, "medium"),  bottom=_lado(_AZUL, "medium"),
)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, color=_AZUL, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------

def generar_reporte(resultados: list[dict], carpeta: str) -> str:
    """
    Crea Resultados_Evaluacion.xlsx en la carpeta indicada.
    Devuelve la ruta completa del archivo generado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.sheet_view.showGridLines = False

    # -----------------------------------------------------------------------
    # Título superior
    # -----------------------------------------------------------------------
    # Determinar dinámicamente cuántas hojas evaluó (por el tamaño de la lista `puntos_hoja` del primero)
    num_hojas = len(resultados[0]["puntos_hoja"]) if resultados else 5
    # El ancho será 2 columnas base (Nombre, Nota) + num_hojas + Observaciones = num_hojas + 3
    fin_col_letra = get_column_letter(num_hojas + 3)
    ws.merge_cells(f"A1:{fin_col_letra}1")
    titulo = ws["A1"]
    titulo.value = f"📊 Resultados — Trabajo Cotidiano  |  {date.today().strftime('%d/%m/%Y')}"
    titulo.font      = Font(bold=True, color=_BLANCO, size=14, name="Calibri")
    titulo.fill      = _fill(_AZUL)
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # -----------------------------------------------------------------------
    # Encabezados de columna (fila 2)
    # -----------------------------------------------------------------------
    headers = (
        ["Nombre del Estudiante"] +
        [f"Hoja {i+1}" for i in range(num_hojas)] +
        ["Nota Final (0-100)", "Observaciones"]
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = Font(bold=True, color=_BLANCO, size=11, name="Calibri")
        cell.fill      = _fill("3A7EC4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDE_CABECERA
    ws.row_dimensions[2].height = 30

    # -----------------------------------------------------------------------
    # Datos de estudiantes
    # -----------------------------------------------------------------------
    for row_idx, est in enumerate(resultados, 3):
        # Nombre
        c_nombre = ws.cell(row=row_idx, column=1, value=est["nombre"])
        c_nombre.font      = Font(name="Calibri", size=11, bold=True)
        c_nombre.alignment = Alignment(vertical="center")
        c_nombre.fill      = _fill(_GRIS) if row_idx % 2 == 0 else _fill(_BLANCO)
        c_nombre.border    = BORDE_NORMAL

        # Puntos por hoja (columnas 2 a 1+num_hojas)
        for hoja_idx in range(num_hojas):
            pts = est["puntos_hoja"][hoja_idx]
            col = hoja_idx + 2
            cell = ws.cell(row=row_idx, column=col, value=pts)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDE_NORMAL
            cell.fill      = _fill(_GRIS) if row_idx % 2 == 0 else _fill(_BLANCO)
            if pts is not None:
                cell.font = Font(name="Calibri", size=11, bold=True,
                                 color=_VERDE if pts == 3 else
                                       _AMARIL if pts == 2 else
                                       _ROJO   if pts == 1 else "666666")

        # Nota final
        nota = est["nota_final"]
        col_nota = num_hojas + 2
        c_nota = ws.cell(row=row_idx, column=col_nota, value=nota)
        c_nota.alignment = Alignment(horizontal="center", vertical="center")
        c_nota.border    = BORDE_NORMAL
        nota_color = _VERDE if nota >= 70 else _AMARIL if nota >= 60 else _ROJO
        c_nota.fill = _fill(nota_color)
        c_nota.font = Font(name="Calibri", size=12, bold=True, color=_BLANCO)

        # Observaciones
        col_obs = num_hojas + 3
        c_obs = ws.cell(row=row_idx, column=col_obs, value=est.get("observaciones", ""))
        c_obs.alignment = Alignment(vertical="center", wrap_text=True)
        c_obs.border    = BORDE_NORMAL
        c_obs.fill      = _fill(_GRIS) if row_idx % 2 == 0 else _fill(_BLANCO)
        c_obs.font      = Font(name="Calibri", size=10, color="444444")

        # Error si lo hubo
        if est.get("error"):
            c_nombre.value = f"⚠ {est['nombre']}"
            c_nombre.font  = Font(name="Calibri", size=11, bold=True, color=_ROJO)
            c_obs.value    = est["error"]

        ws.row_dimensions[row_idx].height = 22

    # -----------------------------------------------------------------------
    # Fila de promedio
    # -----------------------------------------------------------------------
    n_datos = len(resultados)
    if n_datos > 0:
        fila_prom = n_datos + 3
        ws.cell(row=fila_prom, column=1, value="PROMEDIO DEL GRUPO").font = Font(
            bold=True, color=_BLANCO, name="Calibri", size=11
        )
        ws.cell(row=fila_prom, column=1).fill   = _fill("3A7EC4")
        ws.cell(row=fila_prom, column=1).border = BORDE_NORMAL
        ws.cell(row=fila_prom, column=1).alignment = Alignment(horizontal="center", vertical="center")

        col_nota = num_hojas + 2
        for col in range(2, col_nota + 1):
            letra = get_column_letter(col)
            formula = f"=IFERROR(AVERAGE({letra}3:{letra}{fila_prom-1}),\"\")"
            cell = ws.cell(row=fila_prom, column=col, value=formula)
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDE_NORMAL
            cell.fill      = _fill("3A7EC4")
            cell.font      = Font(bold=True, color=_BLANCO, name="Calibri", size=11)

        ws.cell(row=fila_prom, column=num_hojas+3, value="").border = BORDE_NORMAL
        ws.row_dimensions[fila_prom].height = 24

    # -----------------------------------------------------------------------
    # Ancho de columnas
    # -----------------------------------------------------------------------
    anchos = [30] + [9] * num_hojas + [16, 70]
    for col_idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # Congelar encabezados
    ws.freeze_panes = "A3"

    # -----------------------------------------------------------------------
    # Hoja de leyenda
    # -----------------------------------------------------------------------
    ws_ley = wb.create_sheet("Leyenda")
    ws_ley.sheet_view.showGridLines = False
    leyenda_data = [
        ("ESCALA MEP — Taxonomía de Bloom", None, None),
        ("Puntos", "Nivel", "Rango de logro"),
        (0, "No presenta evidencia", "0%"),
        (1, "No logrado",            "> 0% y ≤ 33%"),
        (2, "En proceso",            "> 33% y ≤ 66%"),
        (3, "Logrado",               "> 66%"),
    ]
    colores_ley = [None, "3A7EC4", "666666", _ROJO, _AMARIL, _VERDE]
    for r_idx, (fila, color) in enumerate(zip(leyenda_data, colores_ley), 1):
        for c_idx, val in enumerate(fila, 1):
            cell = ws_ley.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = BORDE_NORMAL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if color:
                cell.fill = _fill(color)
                cell.font = Font(bold=True, color=_BLANCO, name="Calibri")
            else:
                cell.font = Font(name="Calibri", size=11)
        ws_ley.row_dimensions[r_idx].height = 22

    ws_ley.merge_cells("A1:C1")
    ws_ley["A1"].font = Font(bold=True, color=_BLANCO, size=13, name="Calibri")
    ws_ley["A1"].fill = _fill(_AZUL)
    ws_ley.column_dimensions["A"].width = 10
    ws_ley.column_dimensions["B"].width = 28
    ws_ley.column_dimensions["C"].width = 22

    # -----------------------------------------------------------------------
    # Guardar en la raíz del proyecto
    # -----------------------------------------------------------------------
    ruta_salida = str(RUTA_RESULTADOS)
    wb.save(ruta_salida)
    return ruta_salida
