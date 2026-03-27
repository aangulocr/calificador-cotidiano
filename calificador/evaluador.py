"""
evaluador.py — Motor central de evaluación.

Flujo:
  1. cargar_patron()  → lee PLANTILLA.xlsx y construye la configuración
                        de evaluación por hoja a partir de la hoja _CONFIG.
  2. evaluar_estudiante() → compara el libro del estudiante con el patrón
                            usando sólo los temas activos seleccionados por
                            el profesor.
"""

from __future__ import annotations
import os
import difflib
from pathlib import Path

import openpyxl

from .config import (
    HOJA_CONFIG, ARCHIVO_PATRON, MAX_HOJAS,
    porcentaje_a_bloom, TEMA_NOMBRE, RUTA_PATRON
)
from .comparadores import EVALUADORES


# ---------------------------------------------------------------------------
# Fuzzy sheet-name matching
# ---------------------------------------------------------------------------

import unicodedata

def _encontrar_hoja_fuzzy(
    nombre_esperado: str,
    hojas_disponibles: list[str],
) -> tuple[str | None, str]:
    """
    Busca la hoja más parecida al nombre esperado entre las disponibles.
    Devuelve (nombre_encontrado, metodo) o (None, "").

    Estrategias en orden de prioridad:
      1. Exacto
      2. Normalizado (strip, lower, sin acentos)
      3. El nombre del estudiante CONTIENE el nombre esperado
      4. El nombre esperado CONTIENE el nombre del estudiante
      5. difflib similarity >= 0.40
    """
    # 1. Exacto
    if nombre_esperado in hojas_disponibles:
        return nombre_esperado, "exacto"

    def remove_accents(txt: str) -> str:
        s = unicodedata.normalize('NFD', txt).encode('ascii', 'ignore').decode("utf-8")
        return s.strip().lower()

    esperado_n = remove_accents(nombre_esperado)
    hojas_n    = [remove_accents(h) for h in hojas_disponibles]

    # 2. Normalizado vacio
    if not esperado_n:
        return None, ""

    # 2. Normalizado
    for hoja, hoja_n in zip(hojas_disponibles, hojas_n):
        if hoja_n == esperado_n:
            return hoja, "normalizado"

    # 3. El nombre del estudiante contiene el esperado
    # (ej: «JuanPerez_Datos» contiene «datos»)
    for hoja, hoja_n in zip(hojas_disponibles, hojas_n):
        if esperado_n in hoja_n:
            return hoja, "incluido"

    # 4. El nombre esperado contiene el del estudiante
    # (ej: «Hoja Analisis» del patrón vs «Analisis» del estudiante)
    for hoja, hoja_n in zip(hojas_disponibles, hojas_n):
        if hoja_n in esperado_n and len(hoja_n) >= 3:
            return hoja, "contenido"

    # 5. Similitud difflib
    coincidencias = difflib.get_close_matches(
        esperado_n, hojas_n, n=1, cutoff=0.50
    )
    if coincidencias:
        idx = hojas_n.index(coincidencias[0])
        return hojas_disponibles[idx], "similar"

    return None, ""


# ---------------------------------------------------------------------------
# Cargar patrón
# ---------------------------------------------------------------------------

def cargar_patron(carpeta: str) -> dict:
    """
    Lee PLANTILLA.xlsx de la carpeta indicada.

    Devuelve un dict con:
        {
          "libro":  openpyxl.Workbook  (data_only=False),
          "config": {
              "<nombre_hoja>": [
                  {"tema_id": str, "celdas_evaluar": str, "peso": int},
                  ...
              ],
              ...
          },
          "orden_hojas": [nombre_hoja_1, nombre_hoja_2, ...]
        }

    La hoja _CONFIG debe tener las columnas (fila 1 = encabezados):
        hoja_nombre | hoja_indice | tema_id | celdas_evaluar | peso
    """
    ruta = RUTA_PATRON
    if not ruta.exists():
        raise FileNotFoundError(
            f"No se encontró '{ARCHIVO_PATRON}' en la raíz del proyecto:\n{ruta.parent}"
        )

    wb = openpyxl.load_workbook(str(ruta), data_only=False)

    if HOJA_CONFIG not in wb.sheetnames:
        raise ValueError(
            f"El patrón no tiene la hoja '{HOJA_CONFIG}'.\n"
            "Consulta el README para preparar PLANTILLA.xlsx"
        )

    ws_cfg = wb[HOJA_CONFIG]
    config: dict[str, list] = {}
    orden_hojas: list[str] = []

    # Detectar índices de columnas en fila 1
    encabezados = {
        str(c.value).strip().lower(): c.column - 1
        for c in ws_cfg[1]
        if c.value
    }

    def _get_col(row, encabezados, nombre):
        idx = encabezados.get(nombre)
        return row[idx] if idx is not None and idx < len(row) else None

    for row in ws_cfg.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        hoja_nombre    = str(_get_col(row, encabezados, "hoja_nombre") or "").strip()
        # Normaliza tema_id para evitar errores comunes de los usuarios (ej: "color fuente" -> "color_fuente")
        tema_id        = str(_get_col(row, encabezados, "tema_id") or "").strip().lower().replace(" ", "_").replace("-", "_")
        celdas_evaluar = str(_get_col(row, encabezados, "celdas_evaluar") or "-").strip()
        try:
            peso = int(_get_col(row, encabezados, "peso") or 1)
        except (TypeError, ValueError):
            peso = 1

        if not hoja_nombre or not tema_id:
            continue

        if hoja_nombre not in config:
            config[hoja_nombre] = []
            orden_hojas.append(hoja_nombre)

        config[hoja_nombre].append({
            "tema_id":        tema_id,
            "celdas_evaluar": celdas_evaluar,
            "peso":           peso,
        })

    return {"libro": wb, "config": config, "orden_hojas": orden_hojas}


# ---------------------------------------------------------------------------
# Evaluar un estudiante
# ---------------------------------------------------------------------------

def evaluar_estudiante(
    ruta_archivo: str,
    patron: dict,
    temas_activos: set[str],
    log_callback=None,
) -> dict:
    """
    Compara el libro del estudiante contra el patrón.

    Parámetros:
        ruta_archivo  : ruta absoluta al .xlsx del estudiante.
        patron        : dict devuelto por cargar_patron().
        temas_activos : set de tema_id seleccionados por el profesor.
        log_callback  : función opcional que recibe un str para logging.

    Devuelve dict:
        {
          "nombre":           str,
          "puntos_hoja":      [int]*5   (0-3 por hoja, None si no existe),
          "nota_final":       int,
          "observaciones":    str,
          "error":            str | None
        }
    """
    nombre = Path(ruta_archivo).stem  # Nombre del archivo sin .xlsx

    wb_pat = patron["libro"]
    config  = patron["config"]
    orden   = patron["orden_hojas"]
    num_hojas_patron = len(orden)

    resultado = {
        "nombre":        nombre,
        "puntos_hoja":   [None] * num_hojas_patron,
        "nota_final":    0,
        "observaciones": "",
        "error":         None,
    }

    def log(msg):
        if log_callback:
            log_callback(msg)

    try:
        wb_est = openpyxl.load_workbook(ruta_archivo, data_only=False)
    except Exception as e:
        resultado["error"] = f"No se pudo abrir el archivo: {e}"
        return resultado

    obs_generales: list[str] = []
    brutos_posibles_total = 0
    brutos_obtenidos_total = 0

    for idx, hoja_nombre in enumerate(orden):
        n_hoja = idx + 1

        # Buscar la hoja del estudiante (con tolerancia a nombres similares)
        hoja_est_nombre, metodo = _encontrar_hoja_fuzzy(
            hoja_nombre, wb_est.sheetnames
        )
        if hoja_est_nombre is None:
            log(f"  ⚠ Hoja '{hoja_nombre}' no entregada por el estudiante")
            obs_generales.append(f"Hoja {n_hoja}: no entregada")
            resultado["puntos_hoja"][idx] = 0
            continue

        if metodo != "exacto":
            log(f"  ℹ Hoja '{hoja_nombre}' → match '{hoja_est_nombre}' ({metodo})")

        ws_est = wb_est[hoja_est_nombre]
        hoja_pat_nombre, _ = _encontrar_hoja_fuzzy(hoja_nombre, wb_pat.sheetnames)
        ws_pat = wb_pat[hoja_pat_nombre] if hoja_pat_nombre else None

        if ws_pat is None:
            log(f"  ⚠ Hoja patrón '{hoja_nombre}' no existe en el archivo de respuestas — omitida")
            # If pattern sheet is completely missing physically, we skip scoring it
            continue

        temas_hoja = config.get(hoja_nombre, [])
        # Filtrar sólo los temas activos
        temas_evaluar = [
            t for t in temas_hoja if t["tema_id"] in temas_activos
        ]

        if not temas_evaluar:
            # Hoja configurada pero ningún tema activo para ella
            resultado["puntos_hoja"][idx] = 3  # Se considera lograda
            continue

        puntos_posibles_total  = 0
        puntos_obtenidos_total = 0
        obs_hoja: list[str] = []

        for tema_cfg in temas_evaluar:
            tema_id = tema_cfg["tema_id"]
            fn_eval = EVALUADORES.get(tema_id)
            if fn_eval is None:
                log(f"  ⚠ No hay evaluador para tema '{tema_id}'")
                continue

            meta = {
                "celdas_evaluar": tema_cfg["celdas_evaluar"],
                "tema_nombre":    TEMA_NOMBRE.get(tema_id, tema_id),
            }

            try:
                obtenidos, posibles, obs = fn_eval(ws_est, ws_pat, meta)
            except Exception as exc:
                log(f"  ✗ Error evaluando '{tema_id}' en '{nombre}': {exc}")
                obtenidos, posibles, obs = 0, tema_cfg["peso"], [f"Error interno al evaluar {tema_id}"]

            # Aplicar peso
            factor = tema_cfg["peso"] / max(posibles, 1)
            puntos_posibles_total  += tema_cfg["peso"]
            puntos_obtenidos_total += round(obtenidos * factor)
            obs_hoja.extend(obs)

        # Totales brutos globales
        brutos_posibles_total += puntos_posibles_total
        brutos_obtenidos_total += puntos_obtenidos_total

        # Porcentaje y escala MEP
        if puntos_posibles_total > 0:
            pct = (puntos_obtenidos_total / puntos_posibles_total) * 100
        else:
            pct = 100.0

        bloom = porcentaje_a_bloom(pct)
        resultado["puntos_hoja"][idx] = bloom

        etiqueta = {0: "No presenta", 1: "No logrado", 2: "En proceso", 3: "Logrado"}[bloom]
        log(f"  ✓ Hoja {n_hoja} '{hoja_nombre}': {int(puntos_obtenidos_total)}/{int(puntos_posibles_total)} pts ({pct:.1f}%) → {bloom}/3 ({etiqueta})")

        if obs_hoja:
            obs_txt = f"Hoja {n_hoja}: " + "; ".join(obs_hoja)
            obs_generales.append(obs_txt)

    if obs_generales:
        resultado["observaciones"] = " | ".join(obs_generales)
    else:
        resultado["observaciones"] = "Todo correcto"

    # Calcular nota final promediando las hojas evaluadas
    hojas_evaluadas = [p for p in resultado["puntos_hoja"] if p is not None]
    if hojas_evaluadas:
        puntos_totales = sum(hojas_evaluadas)
        max_puntos = len(hojas_evaluadas) * 3
        promedio_puntos = puntos_totales / len(hojas_evaluadas)
        # Convertir la escala 0-3 a nota 0-100
        # 3 pts = 100, 2 pts = 66.6, 1 pto = 33.3, 0 pts = 0
        resultado["nota_final"] = round((promedio_puntos / 3.0) * 100)
        log(f"  ★ Puntos MEP: {puntos_totales}/{max_puntos}  |  Puntos Brutos: {int(brutos_obtenidos_total)}/{int(brutos_posibles_total)}  |  (Nota Final: {resultado['nota_final']})")
    else:
        resultado["nota_final"] = 0
        log("  ★ Puntos MEP: 0/0  |  Puntos Brutos: 0/0  |  (Nota Final: 0)")

    return resultado


# ---------------------------------------------------------------------------
# Iterar todos los estudiantes en la carpeta
# ---------------------------------------------------------------------------

def calificar_carpeta(
    carpeta: str,
    patron: dict,
    temas_activos: set[str],
    log_callback=None,
) -> list[dict]:
    """
    Itera todos los .xlsx de la carpeta (excepto el patrón y resultados)
    y devuelve la lista de resultados de evaluar_estudiante().
    """
    carpeta_path = Path(carpeta)
    excluir = {ARCHIVO_PATRON.lower(), "resultados_evaluacion.xlsx"}

    archivos = sorted([
        f for f in carpeta_path.glob("*.xlsx")
        if f.name.lower() not in excluir
    ])

    if not archivos:
        raise FileNotFoundError("No se encontraron archivos .xlsx de estudiantes en la carpeta.")

    resultados = []
    for ruta in archivos:
        if log_callback:
            log_callback(f"\n📄 Procesando: {ruta.name}")
        r = evaluar_estudiante(str(ruta), patron, temas_activos, log_callback)
        resultados.append(r)

    return resultados
